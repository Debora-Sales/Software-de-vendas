import customtkinter as ctk
from tkinter import messagebox
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from database import (
    obter_vendas_por_vendedor,
    obter_relatorio_lucro_db,
    obter_historico_vendas_db,
    obter_entregas_pendentes_db,
    atualizar_status_entrega_db,
    obter_tabela_fretes_db,
    salvar_configuracao_frete_db,
    limpar_historico_vendas_db
)

class JanelaRelatorios(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)

        # Configurações Básicas da Janela (O "Básico no início")
        self.title("Relatórios Financeiros - Xô Sujeira")
        self.geometry("950x850")
        self.resizable(False, False)
        self.grab_set()  # Impede interação com a janela de trás
        self.focus()

        self.entradas_frete = {} # Armazena widgets para salvar configurações

        # Cabeçalho Principal
        ctk.CTkLabel(self, text="📊 Gestão Financeira e Resultados", font=("Roboto", 28, "bold")).pack(pady=20)

        # Sistema de Abas (Mesma lógica do Clientes/Produtos)
        self.tabview = ctk.CTkTabview(self, width=900, height=600)
        self.tabview.pack(padx=20, pady=10, fill="both", expand=True)

        self.tab_comissoes = self.tabview.add("Comissões")
        self.tab_lucro = self.tabview.add("Margem de Lucro")
        self.tab_historico = self.tabview.add("Histórico de Vendas")
        self.tab_logistica = self.tabview.add("Gestão de Entregas")

        # Inicialização das interfaces das abas
        self._setup_ui_comissoes()
        self._setup_ui_lucro()
        self._setup_ui_historico()
        self._setup_ui_logistica()

        # --- FRAME DE AÇÕES INFERIORES (Script 43) ---
        self.frame_acoes_finais = ctk.CTkFrame(self, fg_color="transparent")
        self.frame_acoes_finais.pack(pady=25)

        self.btn_pdf = ctk.CTkButton(
            self.frame_acoes_finais,
            text="📄 Exportar PDF",
            height=50,
            fg_color="#A52A2A",
            hover_color="#800000",
            font=("Roboto", 16, "bold"),
            command=self.exportar_pdf
        )
        self.btn_pdf.pack(side="left", padx=10)

        self.btn_atualizar = ctk.CTkButton(
            self.frame_acoes_finais, 
            text="🔄 ATUALIZAR RELATÓRIOS", 
            height=50, 
            fg_color="teal", 
            hover_color="#006d6d",
            font=("Roboto", 16, "bold"),
            command=self.carregar_dados
        )
        self.btn_atualizar.pack(side="left", padx=10)

        self.btn_excel = ctk.CTkButton(
            self.frame_acoes_finais,
            text="Excel 📥",
            height=50,
            fg_color="#2E8B57",
            hover_color="#1E6B47",
            font=("Roboto", 16, "bold"),
            command=self.exportar_excel
        )
        self.btn_excel.pack(side="left", padx=10)

        # Carregamento inicial automático
        self.carregar_dados()

    def _setup_ui_comissoes(self):
        header = ctk.CTkFrame(self.tab_comissoes, fg_color="gray30")
        header.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(header, text="Vendedor", width=350, anchor="w", font=("", 12, "bold")).pack(side="left", padx=20)
        ctk.CTkLabel(header, text="Total Vendido", width=200, font=("", 12, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Comissão (5%)", width=200, font=("", 12, "bold")).pack(side="left")

        self.scroll_comissoes = ctk.CTkScrollableFrame(self.tab_comissoes, width=850, height=450)
        self.scroll_comissoes.pack(padx=10, pady=5, fill="both", expand=True)

    def _setup_ui_lucro(self):
        header = ctk.CTkFrame(self.tab_lucro, fg_color="gray30")
        header.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(header, text="Produto", width=280, anchor="w", font=("", 12, "bold")).pack(side="left", padx=20)
        ctk.CTkLabel(header, text="Qtd", width=60, font=("", 12, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Receita Bruta", width=150, font=("", 12, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Custo Total", width=150, font=("", 12, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Lucro Líquido", width=150, font=("", 12, "bold")).pack(side="left")

        self.scroll_lucro = ctk.CTkScrollableFrame(self.tab_lucro, width=850, height=450)
        self.scroll_lucro.pack(padx=10, pady=5, fill="both", expand=True)

    def _setup_ui_historico(self):
        header = ctk.CTkFrame(self.tab_historico, fg_color="gray30")
        header.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(header, text="ID", width=40, font=("", 11, "bold")).pack(side="left", padx=5)
        ctk.CTkLabel(header, text="Vendedor", width=120, anchor="w", font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Cliente", width=120, anchor="w", font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Data", width=100, font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Pagto", width=80, font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Tipo", width=80, font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Urgência", width=80, font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Status", width=90, font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Total", width=100, font=("", 11, "bold")).pack(side="left")

        self.scroll_historico = ctk.CTkScrollableFrame(self.tab_historico, width=850, height=450)
        self.scroll_historico.pack(padx=10, pady=5, fill="both", expand=True)

        # Script 41: Botão para limpar histórico de vendas
        self.btn_limpar_hist = ctk.CTkButton(self.tab_historico, text="🗑️ Limpar Todo o Histórico", fg_color="red", command=self.confirmar_limpeza_historico)
        self.btn_limpar_hist.pack(pady=10)

    def _setup_ui_logistica(self):
        header = ctk.CTkFrame(self.tab_logistica, fg_color="gray30")
        header.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(header, text="ID", width=40, font=("", 11, "bold")).pack(side="left", padx=5)
        ctk.CTkLabel(header, text="Cliente", width=150, anchor="w", font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Endereço de Entrega", width=250, anchor="w", font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Urgência", width=80, font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Status", width=100, font=("", 11, "bold")).pack(side="left")
        ctk.CTkLabel(header, text="Ações Logísticas", width=200, font=("", 11, "bold")).pack(side="left")

        self.scroll_logistica = ctk.CTkScrollableFrame(self.tab_logistica, width=850, height=450)
        self.scroll_logistica.pack(padx=10, pady=5, fill="both", expand=True)

    def mudar_status(self, id_venda, novo_status):
        """Atualiza o status e recarrega a lista."""
        if atualizar_status_entrega_db(id_venda, novo_status):
            messagebox.showinfo("Logística", f"Venda #{id_venda} marcada como '{novo_status}'.")
            self.carregar_dados()

    def confirmar_limpeza_historico(self):
        """Script 41: Solicita confirmação antes de apagar o banco de vendas."""
        if messagebox.askyesno("Confirmar Limpeza", "⚠️ ATENÇÃO: Deseja realmente apagar TODO o histórico de vendas?\nEsta ação não pode ser desfeita."):
            if limpar_historico_vendas_db():
                messagebox.showinfo("Sucesso", "Histórico de vendas removido com sucesso.")
                self.carregar_dados()

    def carregar_dados(self):
        """Limpa os frames e recarrega os dados do Banco de Dados."""
        # 1. Comissões
        for w in self.scroll_comissoes.winfo_children(): w.destroy()
        for nome, total in obter_vendas_por_vendedor():
            f = ctk.CTkFrame(self.scroll_comissoes)
            f.pack(fill="x", pady=2)
            ctk.CTkLabel(f, text=nome, width=350, anchor="w").pack(side="left", padx=20)
            ctk.CTkLabel(f, text=f"R$ {total:,.2f}", width=200).pack(side="left")
            ctk.CTkLabel(f, text=f"R$ {total*0.05:,.2f}", width=200, text_color="#4CAF50", font=("", 12, "bold")).pack(side="left")

        # 2. Lucratividade (RN04)
        for w in self.scroll_lucro.winfo_children(): w.destroy()
        for nome, qtd, venda, custo in obter_relatorio_lucro_db():
            lucro = (venda or 0) - (custo or 0)
            f = ctk.CTkFrame(self.scroll_lucro)
            f.pack(fill="x", pady=2)
            ctk.CTkLabel(f, text=nome, width=280, anchor="w").pack(side="left", padx=20)
            ctk.CTkLabel(f, text=str(qtd), width=60).pack(side="left")
            ctk.CTkLabel(f, text=f"R$ {venda:,.2f}", width=150).pack(side="left")
            ctk.CTkLabel(f, text=f"R$ {custo:,.2f}", width=150).pack(side="left")
            cor_lucro = "cyan" if lucro >= 0 else "red"
            ctk.CTkLabel(f, text=f"R$ {lucro:,.2f}", width=150, text_color=cor_lucro, font=("", 12, "bold")).pack(side="left")

        # 3. Histórico de Vendas
        for w in self.scroll_historico.winfo_children(): w.destroy()
        dados_historico = obter_historico_vendas_db()
        for id_v, vend, cli, data, total, forma, tipo, urg, status in dados_historico:
            f = ctk.CTkFrame(self.scroll_historico)
            f.pack(fill="x", pady=2)
            ctk.CTkLabel(f, text=str(id_v), width=40).pack(side="left", padx=5)
            ctk.CTkLabel(f, text=vend, width=120, anchor="w").pack(side="left")
            ctk.CTkLabel(f, text=cli, width=120, anchor="w").pack(side="left")
            ctk.CTkLabel(f, text=data, width=100).pack(side="left")
            ctk.CTkLabel(f, text=forma, width=80).pack(side="left")
            
            # Cores para o tipo e urgência para facilitar leitura
            cor_urgencia = "orange" if urg in ["Urgente", "Crítico"] else "gray"
            cor_tipo = "cyan" if tipo == "Entrega" else "gray"
            
            # Lógica para o Status da Entrega (Sucesso)
            texto_status = status
            cor_status = "gray"
            if tipo == "Entrega" and status == "Entregue":
                texto_status = "✅ Sucesso"
                cor_status = "green"
            
            ctk.CTkLabel(f, text=tipo, width=80, text_color=cor_tipo).pack(side="left")
            ctk.CTkLabel(f, text=urg, width=80, text_color=cor_urgencia).pack(side="left")
            ctk.CTkLabel(f, text=texto_status, width=90, text_color=cor_status, font=("", 11, "bold")).pack(side="left")
            ctk.CTkLabel(f, text=f"R$ {total:,.2f}", width=100, text_color="orange", font=("", 11, "bold")).pack(side="left")

        # 4. Gestão de Logística
        for w in self.scroll_logistica.winfo_children(): w.destroy()
        dados_logistica = obter_entregas_pendentes_db()
        for id_v, cli, ender, urg, status in dados_logistica:
            f = ctk.CTkFrame(self.scroll_logistica)
            f.pack(fill="x", pady=2)
            
            ctk.CTkLabel(f, text=str(id_v), width=40).pack(side="left", padx=5)
            ctk.CTkLabel(f, text=cli, width=150, anchor="w").pack(side="left")
            ctk.CTkLabel(f, text=ender, width=250, anchor="w", font=("", 10)).pack(side="left")
            
            cor_urgencia = "red" if urg == "Crítico" else "orange" if urg == "Urgente" else "gray"
            ctk.CTkLabel(f, text=urg, width=80, text_color=cor_urgencia, font=("", 11, "bold")).pack(side="left")
            
            cor_status = "cyan" if status == "Em Rota" else "yellow"
            ctk.CTkLabel(f, text=status, width=100, text_color=cor_status).pack(side="left")

            # Botões de Ação Logística
            ctk.CTkButton(f, text="🚛 Em Rota", width=90, height=24, fg_color="#1f538d", 
                          command=lambda iv=id_v: self.mudar_status(iv, "Em Rota")).pack(side="left", padx=5)
            ctk.CTkButton(f, text="✅ Entregue", width=90, height=24, fg_color="green", 
                          command=lambda iv=id_v: self.mudar_status(iv, "Entregue")).pack(side="left", padx=5)

        if not dados_logistica:
            ctk.CTkLabel(self.scroll_logistica, text="Nenhuma entrega pendente encontrada.", text_color="gray").pack(pady=40)

    def exportar_pdf(self):
        """Gera um PDF estilizado com cara de dashboard/relatório executivo."""
        try:
            dados_brutos = obter_historico_vendas_db()
            if not dados_brutos:
                messagebox.showwarning("Aviso", "Não há dados para exportar.")
                return

            nome_arquivo = "Dashboard_Vendas.pdf"
            doc = SimpleDocTemplate(nome_arquivo, pagesize=landscape(A4))
            elementos = []
            estilos = getSampleStyleSheet()

            # Título e Resumo
            total_vendas = sum(item[4] for item in dados_brutos)
            ticket_medio = total_vendas / len(dados_brutos)

            elementos.append(Paragraph("XÔ SUJEIRA - DASHBOARD EXECUTIVO DE VENDAS", estilos['Title']))
            elementos.append(Spacer(1, 12))
            elementos.append(Paragraph(f"<b>Resumo Financeiro:</b> Total R$ {total_vendas:,.2f} | Ticket Médio: R$ {ticket_medio:,.2f} | Total de Registros: {len(dados_brutos)}", estilos['Normal']))
            elementos.append(Spacer(1, 20))

            # Preparação da Tabela
            header = ["ID", "Vendedor", "Cliente", "Data", "Total (R$)", "Pagto", "Tipo", "Urgência", "Status"]
            dados_tabela = [header]
            
            for item in dados_brutos:
                linha = list(item)
                linha[4] = f"{item[4]:,.2f}" # Formata valor
                dados_tabela.append(linha)

            t = Table(dados_tabela, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.teal),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white])
            ]))

            elementos.append(t)
            doc.build(elementos)
            messagebox.showinfo("Sucesso", f"Dashboard PDF gerado: {nome_arquivo}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao gerar PDF: {e}")

    def exportar_excel(self):
        """Exporta para Excel com um resumo de KPIs estilizado e tabela formatada."""
        try:
            dados_historico = obter_historico_vendas_db()
            if not dados_historico:
                messagebox.showwarning("Aviso", "Não há dados para exportar.")
                return

            colunas = ["ID", "Vendedor", "Cliente", "Data", "Total", "Pagamento", "Tipo", "Urgência", "Status"]
            df = pd.DataFrame(dados_historico, columns=colunas)
            nome_arquivo = "Dashboard_Vendas.xlsx"

            # Cálculos dos indicadores para o Dashboard
            total_vendas = df["Total"].sum()
            ticket_medio = df["Total"].mean()
            total_pedidos = len(df)

            with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
                # Inicia a tabela de dados na linha 8 para deixar o topo livre para o Dashboard
                df.to_excel(writer, sheet_name='Dashboard', startrow=7, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Dashboard']

                # --- CONFIGURAÇÃO DE ESTILOS ---
                estilo_titulo = Font(name='Segoe UI', size=16, bold=True, color="004D4D")
                estilo_header = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
                preenchimento_teal = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
                alinhamento_centro = Alignment(horizontal="center", vertical="center")
                borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                # --- TÍTULO PRINCIPAL ---
                worksheet.merge_cells('A1:I1')
                worksheet['A1'] = "XÔ SUJEIRA - RELATÓRIO GERENCIAL DE VENDAS"
                worksheet['A1'].font = estilo_titulo
                worksheet['A1'].alignment = alinhamento_centro

                # --- CARDS DE INDICADORES (KPIs) ---
                indicadores = [
                    ("A3", "FATURAMENTO TOTAL", f"R$ {total_vendas:,.2f}", "A4"),
                    ("D3", "TICKET MÉDIO", f"R$ {ticket_medio:,.2f}", "D4"),
                    ("G3", "TOTAL DE PEDIDOS", str(total_pedidos), "G4")
                ]

                for pos_label, label, valor, pos_val in indicadores:
                    worksheet[pos_label] = label
                    worksheet[pos_label].font = Font(bold=True, size=10, color="666666")
                    worksheet[pos_label].alignment = alinhamento_centro
                    
                    worksheet[pos_val] = valor
                    worksheet[pos_val].font = Font(bold=True, size=14, color="D2691E") # Cor Chocolate/Laranja
                    worksheet[pos_val].alignment = alinhamento_centro
                    worksheet[pos_val].border = borda_fina

                # --- FORMATAÇÃO DA TABELA DE VENDAS ---
                worksheet['A7'] = "DETALHAMENTO DAS TRANSAÇÕES"
                worksheet['A7'].font = Font(bold=True, color="008080")

                # Estilizar o cabeçalho da tabela (Linha 8)
                for col_num, _ in enumerate(colunas, 1):
                    cell = worksheet.cell(row=8, column=col_num)
                    cell.font = estilo_header
                    cell.fill = preenchimento_teal
                    cell.alignment = alinhamento_centro
                    cell.border = borda_fina

                # Formatação de conteúdo e moeda
                for row in range(9, 9 + len(df)):
                    for col in range(1, len(colunas) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = borda_fina
                        if col == 5: # Coluna 'Total'
                            cell.number_format = 'R$ #,##0.00'
                        if col in [1, 4, 8, 9]: # ID, Data, Urgência e Status
                            cell.alignment = alinhamento_centro

                # --- AUTO-AJUSTE DAS COLUNAS ---
                for column_cells in worksheet.columns:
                    # Usamos a célula da linha 8 (index 7) para obter a letra da coluna, 
                    # pois ela faz parte do cabeçalho e nunca está mesclada.
                    column_letter = column_cells[7].column_letter
                    
                    max_length = 0
                    for cell in column_cells[7:]: # Calcula o tamanho apenas do cabeçalho para baixo
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    worksheet.column_dimensions[column_letter].width = max_length + 4

            messagebox.showinfo("Sucesso", f"Dashboard Excel gerado com sucesso: {nome_arquivo}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao gerar Excel: {e}")